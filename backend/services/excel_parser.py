from __future__ import annotations

import csv
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.services.amount_parser import parse_amount
from backend.services.matter_validator import entity_key


def _normalize(value: Any) -> str:
    return str(value or "").strip().lower()


def _clean(value: Any) -> str | None:
    text = " ".join(str(value or "").split()).strip(" \n\t\r,;")
    return text or None


def _format_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _read_csv_rows(path: Path) -> list[tuple[object, ...]]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                return [tuple(row) for row in reader]
        except UnicodeDecodeError:
            continue
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        reader = csv.reader(handle)
        return [tuple(row) for row in reader]


def _tabular_sources(path: Path) -> list[dict[str, object]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv_rows(path)
        if not rows:
            return [{"title": "CSV Import", "header_row": (), "rows": [], "source_format": "csv"}]
        return [{"title": "CSV Import", "header_row": rows[0], "rows": rows[1:], "source_format": "csv"}]

    workbook = load_workbook(path, data_only=True, read_only=True)
    sources: list[dict[str, object]] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            sources.append({"title": sheet.title, "header_row": (), "rows": [], "source_format": "workbook"})
            continue
        sources.append(
            {
                "title": sheet.title,
                "header_row": rows[0],
                "rows": rows[1:],
                "source_format": "workbook",
            }
        )
    return sources


def parse_invoice_workbook(path: str | Path) -> dict[str, object]:
    path = Path(path)
    line_items: list[dict[str, object]] = []
    warnings: list[str] = []
    worksheet_report: list[dict[str, object]] = []
    selected_worksheets: list[str] = []
    ignored_worksheets: list[str] = []
    tabular_sources = _tabular_sources(path)

    target_headers = {
        "organization": {"organization", "plaintiff", "creditor"},
        "company": {"company", "defendant", "debtor", "customer", "customer name", "account"},
        "owner_name": {"owner name", "guarantor", "guarantor name"},
        "company_email": {"company email", "email"},
        "company_number": {"company number", "phone", "phone number"},
        "company_address": {"company address", "address", "service address"},
        "invoice_number": {"invoice", "invoice #", "invoice number", "inv #", "inv"},
        "invoice_date": {"invoice date", "date"},
        "due_date": {"due date"},
        "original_amount": {"original amount", "amount", "invoice amount", "principal"},
        "balance_due": {"balance", "balance due", "amount due", "open amount", "total due", "remaining amount", "remaining balance", "remaining balance due"},
        "county": {"county"},
    }

    debtor_entities: dict[str, str] = {}
    plaintiff_entities: dict[str, str] = {}
    guarantors: dict[str, str] = {}
    conflicts: list[str] = []
    total_original = Decimal("0.00")
    total_balance = Decimal("0.00")
    excluded_low_confidence_rows = 0
    debtor_stats: dict[str, dict[str, object]] = {}

    for source in tabular_sources:
        sheet_warnings: list[str] = []
        sheet_debtor_entities: dict[str, str] = {}
        sheet_plaintiff_entities: dict[str, str] = {}
        sheet_title = str(source.get("title") or "Tabular Import")
        source_format = str(source.get("source_format") or "workbook")
        header_row = tuple(source.get("header_row") or ())
        rows = list(source.get("rows") or [])
        if not header_row:
            worksheet_report.append(
                {
                    "title": sheet_title,
                    "selected": False,
                    "selection_reason": "Tabular source is empty.",
                    "matched_headers": [],
                    "row_count": 0,
                    "parsed_row_count": 0,
                    "usable_row_count": 0,
                    "excluded_low_confidence_rows": 0,
                    "source_format": source_format,
                    "debtor_entities": [],
                    "plaintiff_entities": [],
                    "warnings": [],
                }
            )
            ignored_worksheets.append(sheet_title)
            continue

        normalized_headers = [_normalize(cell) for cell in header_row]
        header_index: dict[str, int] = {}
        for field_name, aliases in target_headers.items():
            for idx, cell in enumerate(normalized_headers):
                if cell in aliases and field_name not in header_index:
                    header_index[field_name] = idx

        sheet_row_count = 0
        sheet_parsed_row_count = 0
        sheet_usable_row_count = 0
        sheet_excluded_low_confidence_rows = 0

        if "balance_due" not in header_index and "original_amount" not in header_index:
            message = f"Sheet '{sheet_title}' did not expose invoice amount headers."
            warnings.append(message)
            sheet_warnings.append(message)
            worksheet_report.append(
                {
                    "title": sheet_title,
                    "selected": False,
                    "selection_reason": "No invoice amount headers were found.",
                    "matched_headers": sorted(header_index.keys()),
                    "row_count": 0,
                    "parsed_row_count": 0,
                    "usable_row_count": 0,
                    "excluded_low_confidence_rows": 0,
                    "source_format": source_format,
                    "debtor_entities": [],
                    "plaintiff_entities": [],
                    "warnings": sheet_warnings,
                }
            )
            ignored_worksheets.append(sheet_title)
            continue

        for row_number, row in enumerate(rows, start=2):
            if not row or not any(value not in (None, "") for value in row):
                continue
            sheet_row_count += 1

            company = _clean(row[header_index["company"]]) if "company" in header_index else None
            organization = _clean(row[header_index["organization"]]) if "organization" in header_index else None
            owner_name = _clean(row[header_index["owner_name"]]) if "owner_name" in header_index else None

            if company:
                company_key = entity_key(company)
                debtor_entities.setdefault(company_key, company)
                sheet_debtor_entities.setdefault(company_key, company)
                if company_key:
                    stats = debtor_stats.setdefault(
                        company_key,
                        {
                            "entity_name": company,
                            "row_count": 0,
                            "usable_row_count": 0,
                            "total_balance_due": Decimal("0.00"),
                            "worksheets": set(),
                        },
                    )
                    stats["row_count"] = int(stats.get("row_count") or 0) + 1
                    stats["worksheets"].add(sheet_title)
                    if len(company) > len(str(stats.get("entity_name") or "")):
                        stats["entity_name"] = company
            if organization:
                plaintiff_entities.setdefault(entity_key(organization), organization)
                sheet_plaintiff_entities.setdefault(entity_key(organization), organization)
            if owner_name:
                guarantors.setdefault(entity_key(owner_name), owner_name)

            raw_original = row[header_index["original_amount"]] if "original_amount" in header_index else None
            raw_balance = row[header_index["balance_due"]] if "balance_due" in header_index else raw_original

            original = parse_amount(raw_original)
            balance = parse_amount(raw_balance)
            parse_warning = balance.warning or original.warning

            if not original.value and not balance.value:
                if parse_warning:
                    message = f"{path.name} {sheet_title}!row {row_number}: {parse_warning}"
                    warnings.append(message)
                    sheet_warnings.append(message)
                    excluded_low_confidence_rows += 1
                    sheet_excluded_low_confidence_rows += 1
                continue

            sheet_parsed_row_count += 1

            parse_confidence = min(original.confidence or 0.0, balance.confidence or 0.0) if raw_original not in (None, "") else balance.confidence
            included_in_total = bool(balance.usable)

            if parse_warning:
                message = f"{path.name} {sheet_title}!row {row_number}: {parse_warning}"
                warnings.append(message)
                sheet_warnings.append(message)
            if not included_in_total:
                excluded_low_confidence_rows += 1
                sheet_excluded_low_confidence_rows += 1

            original_amount = original.value if original.value is not None else Decimal("0.00")
            balance_amount = balance.value if balance.value is not None else Decimal("0.00")
            if included_in_total:
                total_original += original_amount
                total_balance += balance_amount
                sheet_usable_row_count += 1
                if company:
                    company_key = entity_key(company)
                    if company_key and company_key in debtor_stats:
                        stats = debtor_stats[company_key]
                        stats["usable_row_count"] = int(stats.get("usable_row_count") or 0) + 1
                        stats["total_balance_due"] = Decimal(str(stats.get("total_balance_due") or "0.00")) + balance_amount

            item = {
                "organization": organization,
                "company": company,
                "owner_name": owner_name,
                "company_email": _clean(row[header_index["company_email"]]) if "company_email" in header_index else None,
                "company_number": _clean(row[header_index["company_number"]]) if "company_number" in header_index else None,
                "company_address": _clean(row[header_index["company_address"]]) if "company_address" in header_index else None,
                "county": _clean(row[header_index["county"]]) if "county" in header_index else None,
                "invoice_number": _clean(row[header_index["invoice_number"]]) if "invoice_number" in header_index else None,
                "invoice_date": _format_date(row[header_index["invoice_date"]]) if "invoice_date" in header_index else None,
                "due_date": _format_date(row[header_index["due_date"]]) if "due_date" in header_index else None,
                "customer_name": company,
                "original_amount": str(original_amount),
                "balance_due": str(balance_amount),
                "raw_original_amount_text": original.raw_text,
                "raw_balance_due_text": balance.raw_text,
                "parse_confidence": float(parse_confidence or 0.0),
                "parse_warning": parse_warning,
                "included_in_total": included_in_total,
                "currency": "USD",
                "source_sheet": sheet_title,
                "source_row": row_number,
            }
            line_items.append(item)

        if len(sheet_debtor_entities) > 1:
            message = (
                f"Worksheet '{sheet_title}' contains multiple debtor companies. "
                "DocketMint will require debtor selection or split handling before treating it as a single matter."
            )
            warnings.append(message)
            sheet_warnings.append(message)

        selected = sheet_parsed_row_count > 0
        selection_reason = "Worksheet contains usable aging rows." if selected else "No usable invoice rows were found."
        worksheet_report.append(
            {
                "title": sheet_title,
                "selected": selected,
                "selection_reason": selection_reason,
                "matched_headers": sorted(header_index.keys()),
                "row_count": sheet_row_count,
                "parsed_row_count": sheet_parsed_row_count,
                "usable_row_count": sheet_usable_row_count,
                "excluded_low_confidence_rows": sheet_excluded_low_confidence_rows,
                "source_format": source_format,
                "debtor_entities": sorted([value for value in sheet_debtor_entities.values() if value]),
                "plaintiff_entities": sorted([value for value in sheet_plaintiff_entities.values() if value]),
                "warnings": sheet_warnings,
            }
        )
        if selected:
            selected_worksheets.append(sheet_title)
        else:
            ignored_worksheets.append(sheet_title)

    seen_by_invoice: dict[str, str] = {}
    for item in line_items:
        invoice_number = str(item.get("invoice_number") or "").strip().upper()
        if not invoice_number or not item.get("included_in_total"):
            continue
        current_balance = str(item.get("balance_due"))
        previous = seen_by_invoice.get(invoice_number)
        if previous is not None and previous != current_balance:
            conflicts.append(
                f"Invoice {invoice_number} appears with conflicting balances ({previous} vs {current_balance}) in workbook '{path.name}'."
            )
        else:
            seen_by_invoice[invoice_number] = current_balance

    multi_entity_conflict = len([value for value in debtor_entities.values() if value]) > 1
    selection_candidates = [
        {
            "entity_key": key,
            "entity_name": str(stats.get("entity_name") or ""),
            "row_count": int(stats.get("row_count") or 0),
            "usable_row_count": int(stats.get("usable_row_count") or 0),
            "total_balance_due": str(Decimal(str(stats.get("total_balance_due") or "0.00"))),
            "worksheets": sorted(str(item) for item in (stats.get("worksheets") or set()) if item),
        }
        for key, stats in sorted(
            debtor_stats.items(),
            key=lambda item: (
                -int(item[1].get("usable_row_count") or 0),
                -Decimal(str(item[1].get("total_balance_due") or "0.00")),
                str(item[1].get("entity_name") or "").lower(),
            ),
        )
        if str(stats.get("entity_name") or "").strip()
    ]
    parser_report = {
        "worksheet_count": len(tabular_sources),
        "selected_worksheet_count": len(selected_worksheets),
        "ignored_worksheet_count": len(ignored_worksheets),
        "selected_worksheets": list(selected_worksheets),
        "ignored_worksheets": list(ignored_worksheets),
        "selected_row_count": len(line_items),
        "selection_candidate_count": len(selection_candidates),
        "multi_entity_conflict": multi_entity_conflict,
        "source_format": "csv" if path.suffix.lower() == ".csv" else "workbook",
        "status": "selection_required" if multi_entity_conflict else ("ready" if line_items else "empty"),
    }

    return {
        "source_filename": path.name,
        "invoice_count": len(line_items),
        "total_original_amount": str(total_original),
        "total_balance_due": str(total_balance),
        "math_verified": len([item for item in line_items if item.get("included_in_total")]) > 0,
        "warnings": warnings,
        "conflicts": conflicts,
        "line_items": line_items,
        "debtor_entities": sorted([value for value in debtor_entities.values() if value]),
        "plaintiff_entities": sorted([value for value in plaintiff_entities.values() if value]),
        "guarantor_entities": sorted([value for value in guarantors.values() if value]),
        "excluded_low_confidence_rows": excluded_low_confidence_rows,
        "worksheet_report": worksheet_report,
        "selected_worksheets": selected_worksheets,
        "ignored_worksheets": ignored_worksheets,
        "selection_candidates": selection_candidates,
        "parser_report": parser_report,
    }
